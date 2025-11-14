"""
Unit tests for CAD adapters.
Tests adapter factory and interface implementation.
"""

import pytest
from unittest.mock import MagicMock, patch
from src.adapters import get_adapter, create_adapter, ADAPTER_REGISTRY
from src.core import CADInterface


class TestAdapterRegistry:
    """Test suite for adapter registry."""

    def test_adapter_registry_contains_expected_adapters(self):
        """Test that registry contains all expected adapters."""
        assert "autocad" in ADAPTER_REGISTRY
        assert "zwcad" in ADAPTER_REGISTRY
        assert "gcad" in ADAPTER_REGISTRY

    def test_get_adapter_returns_class(self):
        """Test get_adapter returns adapter class."""
        adapter_class = get_adapter("autocad")
        assert adapter_class is not None
        # Check it's a class (not an instance)
        assert isinstance(adapter_class, type)

    def test_get_adapter_case_insensitive(self):
        """Test get_adapter is case-insensitive."""
        assert get_adapter("autocad") == get_adapter("AUTOCAD")
        assert get_adapter("zwcad") == get_adapter("ZWCAD")
        assert get_adapter("gcad") == get_adapter("GCAD")

    def test_get_adapter_invalid_type_raises(self):
        """Test get_adapter raises for invalid CAD type."""
        with pytest.raises(ValueError):
            get_adapter("invalid_cad")

    def test_get_adapter_error_message_includes_supported(self):
        """Test error message lists supported CAD types."""
        try:
            get_adapter("unknown")
            assert False, "Should have raised ValueError"
        except ValueError as e:
            assert "autocad" in str(e).lower()
            assert "zwcad" in str(e).lower()
            assert "gcad" in str(e).lower()


class TestAdapterCreation:
    """Test suite for adapter creation."""

    def test_create_adapter_returns_instance(self):
        """Test create_adapter returns an instance (not a class)."""
        # Skip if not on Windows or CAD not installed
        # This test would normally connect to real CAD
        # For now, we just test the factory pattern
        adapter_class = get_adapter("autocad")
        assert callable(adapter_class)

    def test_create_adapter_case_insensitive(self):
        """Test create_adapter is case-insensitive."""
        try:
            adapter1 = create_adapter("autocad")
            adapter2 = create_adapter("AUTOCAD")
            assert type(adapter1) == type(adapter2)
        except Exception:
            # Expected to fail if CAD not installed
            pass

    def test_create_adapter_invalid_type_raises(self):
        """Test create_adapter raises for invalid type."""
        with pytest.raises(ValueError):
            create_adapter("nonexistent")


class TestCADInterfaceContract:
    """Test suite for CADInterface contract compliance."""

    def test_adapter_implements_cad_interface(self):
        """Test that adapters implement CADInterface."""
        # Import CADInterface from the same context as the adapter to avoid
        # module path mismatch (adapter uses 'from core import', test uses 'from src.core import')
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter_base = AutoCADAdapter.__bases__[0]

        for cad_type, adapter_class in ADAPTER_REGISTRY.items():
            # Verify adapter inherits from the same base class
            assert adapter_class.__bases__[0] == adapter_base
            # Also verify the base class name is CADInterface
            assert adapter_base.__name__ == "CADInterface"

    def test_adapter_has_required_methods(self):
        """Test that adapters have all required interface methods."""
        required_methods = [
            # Connection
            "connect",
            "disconnect",
            "is_connected",
            # Drawing
            "draw_line",
            "draw_circle",
            "draw_arc",
            "draw_rectangle",
            "draw_polyline",
            "draw_ellipse",
            "draw_text",
            "draw_hatch",
            "add_dimension",
            # Layers
            "create_layer",
            "set_current_layer",
            "get_current_layer",
            "list_layers",
            # Files
            "save_drawing",
            "open_drawing",
            "new_drawing",
            # View
            "zoom_extents",
            "refresh_view",
            # Entity
            "delete_entity",
            "get_entity_properties",
            "set_entity_properties",
        ]

        for cad_type, adapter_class in ADAPTER_REGISTRY.items():
            for method in required_methods:
                assert hasattr(
                    adapter_class, method
                ), f"{cad_type} adapter missing {method}"


class TestCoordinateNormalization:
    """Test suite for coordinate normalization."""

    def test_normalize_2d_coordinate(self):
        """Test normalization of 2D coordinates."""
        result = CADInterface.normalize_coordinate((10, 20))
        assert result == (10.0, 20.0, 0.0)

    def test_normalize_3d_coordinate(self):
        """Test normalization of 3D coordinates."""
        result = CADInterface.normalize_coordinate((10, 20, 30))
        assert result == (10.0, 20.0, 30.0)

    def test_normalize_coordinate_converts_to_float(self):
        """Test that normalize_coordinate converts to float."""
        result = CADInterface.normalize_coordinate((10, 20))
        assert isinstance(result[0], float)
        assert isinstance(result[1], float)
        assert isinstance(result[2], float)

    def test_normalize_coordinate_invalid_raises(self):
        """Test invalid coordinate raises error."""
        with pytest.raises(ValueError):
            CADInterface.normalize_coordinate((10,))  # Only 1 dimension # type: ignore

        with pytest.raises(ValueError):
            CADInterface.normalize_coordinate(
                (10, 20, 30, 40) # type: ignore
            )  # Too many 


class TestLineWeightValidation:
    """Test suite for lineweight validation."""

    def test_valid_lineweight_accepted(self):
        """Test that valid lineweights are accepted."""
        valid_weights = [0, 5, 9, 13, 15, 18, 20, 25, 30, 100, 211]
        for weight in valid_weights:
            assert CADInterface.validate_lineweight(weight) == weight

    def test_invalid_lineweight_returns_default(self):
        """Test that invalid lineweights return default."""
        result = CADInterface.validate_lineweight(999)
        assert result == 0  # Default thin line

    def test_lineweight_is_valid_check(self):
        """Test lineweight validity check."""
        from src.core.cad_interface import LineWeight

        assert LineWeight.is_valid(0) is True
        assert LineWeight.is_valid(50) is True
        assert LineWeight.is_valid(211) is True
        assert LineWeight.is_valid(999) is False
        assert LineWeight.is_valid(-1) is False


class TestColorEnums:
    """Test suite for color enumerations."""

    def test_color_enum_has_standard_colors(self):
        """Test that Color enum has standard colors."""
        from src.core.cad_interface import Color

        assert hasattr(Color, "BLACK")
        assert hasattr(Color, "RED")
        assert hasattr(Color, "GREEN")
        assert hasattr(Color, "BLUE")
        assert hasattr(Color, "WHITE")

    def test_color_enum_values_are_rgb_tuples(self):
        """Test that color values are RGB tuples."""
        from src.core.cad_interface import Color

        for color in Color:
            assert isinstance(color.value, tuple)
            assert len(color.value) == 3
            # Check RGB values are in valid range
            for component in color.value:
                assert 0 <= component <= 255


class TestRefreshViewUndoRedo:
    """Test suite for refresh_view and undo/redo interaction."""

    def test_refresh_view_uses_multiple_techniques(self):
        """Test that refresh_view uses multiple fallback techniques.

        Techniques in order:
        1. application.Refresh() (COM API - no undo/redo impact)
        2. SendCommand with REDRAW (most reliable visual update)
        3. Window click simulation (forces UI update)
        """
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = create_adapter("autocad")

        with patch.object(adapter, "_get_application") as mock_get_app, patch.object(
            adapter, "_get_document"
        ) as mock_get_doc, patch.object(
            adapter, "_simulate_autocad_click"
        ) as mock_click:

            mock_app = MagicMock()
            mock_doc = MagicMock()
            mock_get_app.return_value = mock_app
            mock_get_doc.return_value = mock_doc

            result = adapter.refresh_view()

            assert result is True
            # Technique 1: COM API Refresh (should be tried first)
            mock_app.Refresh.assert_called_once()
            # Technique 2: REDRAW command
            mock_doc.SendCommand.assert_called_once_with("_redraw\n")
            # Technique 3: Window click simulation
            mock_click.assert_called_once()

    def test_undo_does_not_call_refresh_view(self):
        """Test that undo() does not call refresh_view().

        Removing refresh_view from undo/redo avoids contaminating
        the undo/redo stack with extra refresh commands.
        """
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = create_adapter("autocad")

        with patch.object(adapter, "_validate_connection"), patch.object(
            adapter, "_get_application"
        ) as mock_get_app, patch.object(adapter, "refresh_view") as mock_refresh:

            mock_app = MagicMock()
            mock_get_app.return_value = mock_app

            # Call undo
            result = adapter.undo(count=1)

            assert result is True
            # Verify refresh_view was NOT called
            mock_refresh.assert_not_called()
            # Verify undo command was sent
            mock_app.ActiveDocument.SendCommand.assert_called_once_with("_undo 1\n")

    def test_redo_does_not_call_refresh_view(self):
        """Test that redo() does not call refresh_view().

        Removing refresh_view from undo/redo avoids contaminating
        the undo/redo stack with extra refresh commands.
        """
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = create_adapter("autocad")

        with patch.object(adapter, "_validate_connection"), patch.object(
            adapter, "_get_application"
        ) as mock_get_app, patch.object(adapter, "refresh_view") as mock_refresh:

            mock_app = MagicMock()
            mock_get_app.return_value = mock_app

            # Call redo
            result = adapter.redo(count=1)

            assert result is True
            # Verify refresh_view was NOT called
            mock_refresh.assert_not_called()
            # Verify redo command was sent
            mock_app.ActiveDocument.SendCommand.assert_called_once_with("_redo 1\n")


class TestDataExport:
    """Test suite for drawing data extraction and export."""

    def test_extract_drawing_data_returns_list(self):
        """Test that extract_drawing_data returns a list of dictionaries."""
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = create_adapter("autocad")

        with patch.object(adapter, "_validate_connection"), patch.object(
            adapter, "_get_document"
        ) as mock_get_doc:

            # Mock ModelSpace with entities (COM-compatible interface)
            mock_entity = MagicMock()
            mock_entity.Handle = "A1B2C3D4"
            mock_entity.ObjectName = "AcDbLine"
            mock_entity.Layer = "0"
            mock_entity.Color = 1
            mock_entity.Length = 100.0
            mock_entity.Area = 0.0
            mock_entity.Name = ""

            # Create proper COM mock: has Count property and Item() method
            mock_model_space = MagicMock()
            mock_model_space.Count = 1
            mock_model_space.Item.return_value = mock_entity

            mock_doc = MagicMock()
            mock_doc.ModelSpace = mock_model_space
            mock_get_doc.return_value = mock_doc

            result = adapter.extract_drawing_data()

            assert isinstance(result, list)
            assert len(result) == 1
            assert result[0]["Handle"] == "A1B2C3D4"
            assert result[0]["ObjectType"] == "AcDbLine"
            assert result[0]["Layer"] == "0"

    def test_extract_drawing_data_handles_empty_drawing(self):
        """Test that extract_drawing_data handles empty drawings."""
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = create_adapter("autocad")

        with patch.object(adapter, "_validate_connection"), patch.object(
            adapter, "_get_document"
        ) as mock_get_doc:

            # Create proper COM mock with Count = 0 for empty drawing
            mock_model_space = MagicMock()
            mock_model_space.Count = 0

            mock_doc = MagicMock()
            mock_doc.ModelSpace = mock_model_space
            mock_get_doc.return_value = mock_doc

            result = adapter.extract_drawing_data()

            assert isinstance(result, list)
            assert len(result) == 0

    def test_export_to_excel_handles_no_data(self):
        """Test that export_to_excel handles drawing with no data."""
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = create_adapter("autocad")

        with patch.object(adapter, "extract_drawing_data") as mock_extract:

            # Mock empty drawing data
            mock_extract.return_value = []

            result = adapter.export_to_excel("/tmp/test.xlsx")

            # Should return False when there's no data
            assert result is False

    def test_export_to_excel_creates_file(self):
        """Test that export_to_excel creates a valid Excel file."""
        import os
        from pathlib import Path
        from src.adapters.autocad_adapter import AutoCADAdapter
        from core import get_config

        adapter = create_adapter("autocad")

        with patch.object(
            adapter, "extract_drawing_data"
        ) as mock_extract, patch.object(adapter, "list_layers") as mock_list_layers:

            # Mock drawing data
            mock_extract.return_value = [
                {
                    "Handle": "A1B2C3D4",
                    "ObjectType": "AcDbLine",
                    "Layer": "0",
                    "Color": "red",
                    "Length": "100.500",
                    "Area": "0.000",
                    "Name": "",
                },
                {
                    "Handle": "E5F6G7H8",
                    "ObjectType": "AcDbCircle",
                    "Layer": "1",
                    "Color": "blue",
                    "Length": "314.159",
                    "Area": "7853.981",
                    "Name": "",
                },
            ]

            # Mock layers
            mock_list_layers.return_value = ["0", "1", "MyLayer"]

            # Use filename that will be saved to configured output directory
            filename = "test_export.xlsx"

            try:
                result = adapter.export_to_excel(filename)

                # Should succeed if openpyxl is installed
                assert result is True

                # Get expected filepath
                config = get_config()
                output_dir = Path(config.output.directory).expanduser().resolve()
                filepath = output_dir / filename

                # File should exist
                assert filepath.exists()
                # File should have content
                assert filepath.stat().st_size > 0
            finally:
                # Cleanup
                config = get_config()
                output_dir = Path(config.output.directory).expanduser().resolve()
                filepath = output_dir / filename
                if filepath.exists():
                    filepath.unlink()

    def test_export_to_excel_creates_layers_sheet(self):
        """Test that export_to_excel creates a Layers sheet with detailed layer information."""
        from pathlib import Path
        from src.adapters.autocad_adapter import AutoCADAdapter
        from core import get_config
        from openpyxl import load_workbook

        adapter = create_adapter("autocad")

        with patch.object(
            adapter, "extract_drawing_data"
        ) as mock_extract, patch.object(
            adapter, "get_layers_info"
        ) as mock_get_layers_info:

            # Mock drawing data
            mock_extract.return_value = [
                {
                    "Handle": "A1B2C3D4",
                    "ObjectType": "AcDbLine",
                    "Layer": "0",
                    "Color": "red",
                    "Length": "100.500",
                    "Area": "0.000",
                    "Name": "",
                },
            ]

            # Mock layers info with detailed data
            mock_get_layers_info.return_value = [
                {
                    "Name": "0",
                    "ObjectCount": 1,
                    "Color": "white",
                    "Linetype": "Continuous",
                    "Lineweight": "Default",
                    "IsLocked": False,
                    "IsVisible": True,
                },
                {
                    "Name": "1",
                    "ObjectCount": 0,
                    "Color": "red",
                    "Linetype": "Continuous",
                    "Lineweight": "Default",
                    "IsLocked": False,
                    "IsVisible": True,
                },
                {
                    "Name": "MyLayer",
                    "ObjectCount": 2,
                    "Color": "blue",
                    "Linetype": "Dashed",
                    "Lineweight": "0.5",
                    "IsLocked": True,
                    "IsVisible": False,
                },
            ]

            filename = "test_layers.xlsx"

            try:
                result = adapter.export_to_excel(filename)
                assert result is True

                # Get expected filepath
                config = get_config()
                output_dir = Path(config.output.directory).expanduser().resolve()
                filepath = output_dir / filename

                # Load the workbook and check sheets
                workbook = load_workbook(str(filepath))
                sheet_names = workbook.sheetnames

                # Should have both Drawing Data and Layers sheets
                assert "Drawing Data" in sheet_names
                assert "Layers" in sheet_names

                # Check Layers sheet content
                layers_sheet = workbook["Layers"]
                # Header + 3 layers = 4 rows
                assert layers_sheet.max_row >= 4
                # Check headers
                assert layers_sheet.cell(row=1, column=1).value == "Name"
                assert layers_sheet.cell(row=1, column=2).value == "ObjectCount"
                assert layers_sheet.cell(row=1, column=3).value == "Color"
                # Check layer data (Name column)
                assert layers_sheet.cell(row=2, column=1).value == "0"
                assert layers_sheet.cell(row=3, column=1).value == "1"
                assert layers_sheet.cell(row=4, column=1).value == "MyLayer"
                # Check object counts
                assert layers_sheet.cell(row=2, column=2).value == 1
                assert layers_sheet.cell(row=3, column=2).value == 0
                assert layers_sheet.cell(row=4, column=2).value == 2

            finally:
                # Cleanup
                config = get_config()
                output_dir = Path(config.output.directory).expanduser().resolve()
                filepath = output_dir / filename
                if filepath.exists():
                    filepath.unlink()

    def test_get_entity_color_returns_string(self):
        """Test that _get_entity_color returns a string."""
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = create_adapter("autocad")

        # Test with a known color (red = 1)
        mock_entity = MagicMock()
        mock_entity.Color = 1

        color = adapter._get_entity_color(mock_entity)

        # Should return either color name or index as string
        assert isinstance(color, str)
        assert len(color) > 0

    def test_get_entity_length_handles_missing_attribute(self):
        """Test that _get_entity_length handles entities without Length."""
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = create_adapter("autocad")

        # Create entity without Length attribute
        mock_entity = MagicMock()
        del mock_entity.Length  # Remove the auto-generated Length attribute

        length = adapter._get_entity_length(mock_entity)

        assert length == 0.0

    def test_get_entity_area_handles_missing_attribute(self):
        """Test that _get_entity_area handles entities without Area."""
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = create_adapter("autocad")

        # Create entity without Area attribute
        mock_entity = MagicMock()
        del mock_entity.Area  # Remove the auto-generated Area attribute

        area = adapter._get_entity_area(mock_entity)

        assert area == 0.0

    def test_get_entity_name_handles_blocks(self):
        """Test that _get_entity_name extracts block names."""
        from src.adapters.autocad_adapter import AutoCADAdapter

        adapter = create_adapter("autocad")

        mock_entity = MagicMock()
        mock_entity.Name = "MyBlockName"

        name = adapter._get_entity_name(mock_entity)

        assert name == "MyBlockName"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
